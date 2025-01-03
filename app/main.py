import os
from string import ascii_uppercase
import openpyxl
from youtube_transcript_api import YouTubeTranscriptApi

def get_video_links_by_sheet(workbook):
    values_by_column = {}

    for sheet in workbook.worksheets:
        for column in sheet.iter_cols():
            column_name = column[0].value
            if column_name not in ["teaser do youtube"]:
                continue
            
            if column_name not in values_by_column.keys():
                values_by_column[column_name] = [cell.value for cell in column]
                continue
            
            values_by_column[column_name] += [cell.value for cell in column]
    
    return values_by_column

def generate_all_register_to_dict_from_worksheet(worksheet):
    columns = []
    registers = []

    for position in range(1, worksheet.max_column + 1):
        columns.append({
            "position": position,
            "name": worksheet.cell(1, position).value
        })
    
    for row in range(2, worksheet.max_row + 1):
        register = {}
        for column in columns:
            value = worksheet.cell(row, column["position"]).value
            register[column["name"]] = value
        
        registers.append(register)

    return registers

def inspect_video(video_link):
    try:
        available_transcripts = []
        
        for part_to_remove in [
            "https://", 
            "www.youtube.com/", 
            "embed/",
        ]:
            video_link = video_link.replace(part_to_remove, "")

        transcript_list = YouTubeTranscriptApi.list_transcripts(video_link)
        for transcript in transcript_list:
            available_transcripts.append({
                "gerada automaticamente": transcript.is_generated,
                "idioma": transcript.language_code,
            })

        return available_transcripts
    except:
        return [{"gerada automaticamente": "N/A", "idioma": "N/A"}]

def complete_register_with_video_info(register):
    if register["teaser do youtube"] == "N/A":
        register["gerada automaticamente"] = "N/A"
        register["idiomas"] = "N/A"

        return register

    transcripts_info = inspect_video(register["teaser do youtube"])

    register["gerada automaticamente"] = ""
    register["idiomas"] = ""
    
    for info in transcripts_info:
        if info["gerada automaticamente"] != "N/A":
            register["gerada automaticamente"] = "sim" if info["gerada automaticamente"] == True else "não"
        else:
            register["gerada automaticamente"] = "N/A"

        if register["idiomas"]:
            register["idiomas"] += f', {info["idioma"]}'
            continue

        register["idiomas"] = info["idioma"]

    return register


def generate_column_letter_position(column_position: int):
    letters = []
    while column_position > 0:
        column_position -= 1
        rest = column_position % len(ascii_uppercase)
        letters.append(ascii_uppercase[rest])
        column_position //= len(ascii_uppercase)
    letter = "".join(letters.__reversed__())

    return letter

def generate_complete_file(complete_registers, file_name):
    columns = list(complete_registers[0].keys())
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for column in columns:
        position = columns.index(column) + 1
        letter = generate_column_letter_position(position)
        worksheet[f"{letter}1"] = column

        for register in complete_registers:
            position = complete_registers.index(register) + 2
            worksheet[f"{letter}{position}"] = register[column]
    
    workbook.save(file_name)

def main():
    workbook = openpyxl.load_workbook("file.xlsx")
    worksheet = workbook.active

    registers_to_dict = generate_all_register_to_dict_from_worksheet(worksheet)
    complete_registers = []

    for register in registers_to_dict:
        complete_registers.append(complete_register_with_video_info(register))
    
    try:
        generate_complete_file(complete_registers, "output.xlsx")
    except Exception as e:
        raise e
    finally:
        print("done")
        # os.remove("output.xlsx")

if __name__ == "__main__":
    main()