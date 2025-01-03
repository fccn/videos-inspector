from string import ascii_uppercase
import openpyxl
from youtube_transcript_api import YouTubeTranscriptApi

class InspectVideosService:

    def _generate_all_register_to_dict_from_worksheet(self, worksheet):
        columns = []
        registers = []

        for position in range(1, worksheet.max_column + 1):
            columns.append({
                "position": position,
                "name": worksheet.cell(1, position).value
            })
        
        for row in range(2, worksheet.max_row + 1):
            register = {}
            for column in columns:
                value = worksheet.cell(row, column["position"]).value
                register[column["name"]] = value
            
            registers.append(register)

        return registers

    def _inspect_video(self, video_link):
        try:
            available_transcripts = []
            
            for part_to_remove in [
                "https://", 
                "www.youtube.com/", 
                "embed/",
            ]:
                video_link = video_link.replace(part_to_remove, "")

            transcript_list = YouTubeTranscriptApi.list_transcripts(video_link)
            for transcript in transcript_list:
                available_transcripts.append({
                    "gerada automaticamente": transcript.is_generated,
                    "idioma": transcript.language_code,
                })

            return available_transcripts
        except:
            return [{"gerada automaticamente": "N/A", "idioma": "N/A"}]

    def _complete_register_with_video_info(self, register):
        if register["teaser do youtube"] == "N/A":
            register["gerada automaticamente"] = "N/A"
            register["idiomas"] = "N/A"

            return register

        transcripts_info = self._inspect_video(register["teaser do youtube"])

        register["gerada automaticamente"] = ""
        register["idiomas"] = ""
        
        for info in transcripts_info:
            if info["gerada automaticamente"] != "N/A":
                register["gerada automaticamente"] = "sim" if info["gerada automaticamente"] == True else "não"
            else:
                register["gerada automaticamente"] = "N/A"

            if register["idiomas"]:
                register["idiomas"] += f', {info["idioma"]}'
                continue

            register["idiomas"] = info["idioma"]

        return register


    def _generate_column_letter_position(self, column_position: int):
        letters = []
        while column_position > 0:
            column_position -= 1
            rest = column_position % len(ascii_uppercase)
            letters.append(ascii_uppercase[rest])
            column_position //= len(ascii_uppercase)
        letter = "".join(letters.__reversed__())

        return letter

    def _generate_complete_file(self, complete_registers, file_name):
        columns = list(complete_registers[0].keys())
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        for column in columns:
            position = columns.index(column) + 1
            letter = self._generate_column_letter_position(position)
            worksheet[f"{letter}1"] = column

            for register in complete_registers:
                position = complete_registers.index(register) + 2
                worksheet[f"{letter}{position}"] = register[column]
        
        workbook.save(file_name)

    def handle_steps_to_inspect_videos(self, file_name, output_path):
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active

        registers_to_dict = self._generate_all_register_to_dict_from_worksheet(worksheet)
        complete_registers = []

        for register in registers_to_dict:
            complete_registers.append(self._complete_register_with_video_info(register))

        self._generate_complete_file(complete_registers, output_path)